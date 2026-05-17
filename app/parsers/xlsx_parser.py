from openpyxl import load_workbook
from app.parsers.base import ParsedChunk

class XlsxParser:
    def parse(self, file_path: str, filename: str) -> list[ParsedChunk]:
        wb = load_workbook(file_path, data_only=True)
        chunks = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                text = " | ".join(str(c) if c is not None else "" for c in row)
                if text.strip(" | "):
                    chunks.append(ParsedChunk(
                        text=text,
                        source_file=filename,
                        page_number=None,
                        section_title=f"Sheet: {sheet_name}"
                    ))
        return chunks